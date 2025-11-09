"""Tkinter based desktop application for the lead verifier."""
from __future__ import annotations

import importlib.util
import itertools
import json
import logging
import os
import queue
import threading
from concurrent.futures import Future, ThreadPoolExecutor
from pathlib import Path
import tkinter as tk
from dataclasses import dataclass
from tkinter import filedialog, messagebox, ttk
from typing import Callable, Dict, Iterable, List, Optional

from ..config import ConfigurationError, load_configuration
from ..factory import build_scrapers
from ..io import write_results
from ..models import AggregatedLeadResult, LeadInput
from ..orchestrator import VerificationOrchestrator
from ..scrapers.sample import EchoScraper


LOGGER = logging.getLogger(__name__)


def _clean_value(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    text = str(value).strip()
    return text or None


def load_rows_from_file(path: str | Path) -> List[Dict[str, str]]:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        import csv

        rows: List[Dict[str, str]] = []
        with file_path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                cleaned = {str(key): _clean_value(value) or "" for key, value in row.items() if key}
                if any(value for value in cleaned.values()):
                    rows.append(cleaned)
        return rows

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        if importlib.util.find_spec("openpyxl") is None:
            raise RuntimeError("Reading Excel files requires the 'openpyxl' package")
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        header = [str(cell).strip() for cell in header_row if cell is not None]
        rows: List[Dict[str, str]] = []
        for values in rows_iter:
            record: Dict[str, str] = {}
            for index, column in enumerate(header):
                cell_value = values[index] if index < len(values) else None
                record[column] = _clean_value(cell_value) or ""
            if any(value for value in record.values()):
                rows.append(record)
        return rows

    if suffix == ".xls":
        if importlib.util.find_spec("pandas") is None:
            raise RuntimeError("Reading .xls files requires the 'pandas' package")
        import pandas as pd  # type: ignore

        dataframe = pd.read_excel(file_path)
        dataframe = dataframe.fillna("")
        rows: List[Dict[str, str]] = []
        for _, series in dataframe.iterrows():
            record = {str(column): _clean_value(series[column]) or "" for column in dataframe.columns}
            if any(value for value in record.values()):
                rows.append(record)
        return rows

    raise ValueError(f"Unsupported file type: {file_path.suffix}")


def normalise_lead_row(row: Dict[str, str], mapping: Dict[str, str]) -> LeadInput:
    mapping = {field: column for field, column in mapping.items() if column}

    def resolve(column: Optional[str]) -> Optional[str]:
        if not column:
            return None
        return _clean_value(row.get(column))

    name = resolve(mapping.get("name")) or None
    phone = resolve(mapping.get("phone")) or None
    email = resolve(mapping.get("email")) or None

    metadata: Dict[str, str] = {
        str(key): (_clean_value(value) or "")
        for key, value in row.items()
        if str(key).strip()
    }

    for field, column in mapping.items():
        if field in {"name", "phone", "email"}:
            continue
        value = resolve(column)
        if value is not None:
            metadata[field] = value

    first_name = metadata.get("first_name")
    last_name = metadata.get("last_name")
    if name:
        parts = name.split()
        if parts and not first_name:
            first_name = parts[0]
        if len(parts) > 1 and not last_name:
            last_name = parts[-1]

    if first_name:
        metadata.setdefault("first_name", first_name)
    if last_name:
        metadata.setdefault("last_name", last_name)

    return LeadInput(
        name=name,
        phone=phone,
        email=email,
        first_name=first_name,
        last_name=last_name,
        metadata=metadata,
    )


def run_verification_job(
    orchestrator: VerificationOrchestrator,
    rows: Iterable[Dict[str, str]],
    mapping: Dict[str, str],
    *,
    cancel_event: Optional[threading.Event] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    result_callback: Optional[Callable[[AggregatedLeadResult], None]] = None,
) -> List[AggregatedLeadResult]:
    lead_rows = list(rows)
    leads = [normalise_lead_row(row, mapping) for row in lead_rows]
    total = len(leads)
    aggregated_results: List[AggregatedLeadResult] = []

    if total == 0:
        if progress_callback:
            progress_callback(0, 0)
        return aggregated_results

    for index, lead in enumerate(leads, start=1):
        if cancel_event and cancel_event.is_set():
            break
        aggregated = orchestrator.verify([lead])[0]
        aggregated_results.append(aggregated)
        if result_callback:
            result_callback(aggregated)
        if progress_callback:
            progress_callback(index, total)
        if cancel_event and cancel_event.is_set():
            break

    return aggregated_results


@dataclass
class SourceStyle:
    """Simple structure describing row styling for a source."""

    background: str
    foreground: str = "#000000"


class MappingFrame(ttk.LabelFrame):
    """Widget that lets the user match spreadsheet columns to known fields."""

    FIELD_LABELS = {
        "name": "Full name *",
        "phone": "Phone",
        "email": "Email",
        "company": "Company",
        "city": "City",
        "state": "State",
    }

    def __init__(self, master: tk.Misc) -> None:
        super().__init__(master, text="2. Column mapping")
        self.variables: Dict[str, tk.StringVar] = {}
        self.comboboxes: Dict[str, ttk.Combobox] = {}
        self._build()

    # ------------------------------------------------------------------
    def _build(self) -> None:
        for row_index, (field_key, label) in enumerate(self.FIELD_LABELS.items()):
            ttk.Label(self, text=label).grid(row=row_index, column=0, sticky="w", padx=4, pady=4)
            variable = tk.StringVar()
            combobox = ttk.Combobox(self, textvariable=variable, state="readonly")
            combobox.grid(row=row_index, column=1, sticky="ew", padx=4, pady=4)
            combobox["values"] = ("",)
            self.variables[field_key] = variable
            self.comboboxes[field_key] = combobox
        self.columnconfigure(1, weight=1)

    # ------------------------------------------------------------------
    def update_options(self, columns: Iterable[str]) -> None:
        values = [""] + list(columns)
        for field, combobox in self.comboboxes.items():
            combobox["values"] = values
            current = combobox.get()
            if current not in values:
                combobox.set("")
        # Auto-select exact matches for convenience
        for column in columns:
            lowered = column.lower()
            for field in self.FIELD_LABELS:
                if field in {"company", "city", "state"}:
                    if lowered == field:
                        self.comboboxes[field].set(column)
                elif lowered == field or lowered.replace(" ", "") == field:
                    self.comboboxes[field].set(column)

    # ------------------------------------------------------------------
    def get_mapping(self) -> Dict[str, str]:
        return {field: variable.get() for field, variable in self.variables.items() if variable.get()}


class LeadVerifierApp:
    """Main application window."""

    SOURCE_COLORS = [
        SourceStyle("#E3F2FD"),
        SourceStyle("#FCE4EC"),
        SourceStyle("#E8F5E9"),
        SourceStyle("#FFF3E0"),
        SourceStyle("#EDE7F6"),
    ]

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Lead Verifier")
        self.root.geometry("1024x720")
        self.root.minsize(900, 640)

        self.orchestrator = self._build_orchestrator()
        self.event_queue: "queue.Queue[tuple]" = queue.Queue()
        self._executor = ThreadPoolExecutor(max_workers=1)
        self.current_task: Optional[Future[List[AggregatedLeadResult]]] = None
        self._cancel_event: Optional[threading.Event] = None
        self.loaded_rows: List[Dict[str, str]] = []
        self.result_rows: List[AggregatedLeadResult] = []
        self.source_styles: Dict[str, SourceStyle] = {}
        self._color_cycle = itertools.cycle(self.SOURCE_COLORS)

        self.file_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Idle")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", lambda *_: self.refresh_result_table())

        self._build_layout()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.after(100, self._poll_queue)

    # ------------------------------------------------------------------
    def _build_layout(self) -> None:
        container = ttk.Frame(self.root, padding=12)
        container.pack(fill="both", expand=True)
        container.columnconfigure(0, weight=1)

        self._build_import_section(container)
        self.mapping_frame = MappingFrame(container)
        self.mapping_frame.grid(row=1, column=0, sticky="nsew", pady=(12, 0))

        self._build_progress_section(container)
        self._build_results_section(container)

    # ------------------------------------------------------------------
    def _build_import_section(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="1. Import leads")
        frame.grid(row=0, column=0, sticky="ew")
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Source file").grid(row=0, column=0, padx=4, pady=4, sticky="w")
        ttk.Entry(frame, textvariable=self.file_var).grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        ttk.Button(frame, text="Browse", command=self.browse_file).grid(row=0, column=2, padx=4, pady=4)

        self.sample_tree = ttk.Treeview(frame, columns=("#1", "#2", "#3"), show="headings", height=5)
        self.sample_tree.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=4, pady=(8, 4))
        frame.rowconfigure(1, weight=1)

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.sample_tree.yview)
        self.sample_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=1, column=3, sticky="ns")

    # ------------------------------------------------------------------
    def _build_progress_section(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="3. Verification progress")
        frame.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        frame.columnconfigure(0, weight=1)

        progress = ttk.Progressbar(frame, maximum=100, variable=self.progress_var)
        progress.grid(row=0, column=0, columnspan=3, sticky="ew", padx=4, pady=4)

        ttk.Label(frame, textvariable=self.status_var).grid(row=1, column=0, columnspan=3, sticky="w", padx=4, pady=4)

        ttk.Button(frame, text="Start", command=self.start_verification).grid(row=2, column=0, padx=4, pady=4, sticky="w")
        ttk.Button(frame, text="Cancel", command=self.cancel_verification).grid(row=2, column=1, padx=4, pady=4, sticky="w")
        ttk.Button(frame, text="Clear results", command=self.clear_results).grid(row=2, column=2, padx=4, pady=4, sticky="e")

    # ------------------------------------------------------------------
    def _build_results_section(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="4. Results")
        frame.grid(row=3, column=0, sticky="nsew", pady=(12, 0))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        filter_bar = ttk.Frame(frame)
        filter_bar.grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        filter_bar.columnconfigure(1, weight=1)

        ttk.Label(filter_bar, text="Filter").grid(row=0, column=0, padx=(0, 8))
        ttk.Entry(filter_bar, textvariable=self.filter_var).grid(row=0, column=1, sticky="ew")

        btn_bar = ttk.Frame(filter_bar)
        btn_bar.grid(row=0, column=2, padx=(8, 0))
        ttk.Button(btn_bar, text="Export CSV", command=self.export_csv).pack(side="left", padx=(0, 4))
        ttk.Button(btn_bar, text="Export Excel", command=self.export_excel).pack(side="left")

        columns = ("lead", "contacts", "sources", "metadata")
        self.results_tree = ttk.Treeview(frame, columns=columns, show="headings")
        for column, heading in zip(columns, ["Lead", "Contacts", "Sources", "Metadata"]):
            self.results_tree.heading(column, text=heading)
            self.results_tree.column(column, anchor="w")
        self.results_tree.grid(row=1, column=0, sticky="nsew", padx=4, pady=4)

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=scroll.set)
        scroll.grid(row=1, column=1, sticky="ns")

    # ------------------------------------------------------------------
    def browse_file(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Spreadsheets", "*.csv *.xlsx *.xls"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xls"), ("All files", "*.*")])
        if not path:
            return
        try:
            rows = load_rows_from_file(path)
        except Exception as exc:  # pragma: no cover - GUI surface
            messagebox.showerror("Import failed", str(exc))
            return
        if not rows:
            messagebox.showwarning("No rows", "The selected file did not contain any leads.")
            return

        self.file_var.set(path)
        self.loaded_rows = rows
        columns = list(rows[0].keys())
        self.mapping_frame.update_options(columns)
        self._populate_sample_tree(columns, rows[:5])
        self.status_var.set(f"Loaded {len(rows)} leads")

    # ------------------------------------------------------------------
    def _populate_sample_tree(self, columns: List[str], rows: List[Dict[str, str]]) -> None:
        self.sample_tree.delete(*self.sample_tree.get_children())
        self.sample_tree["columns"] = columns[:3] if len(columns) > 3 else columns
        for column in self.sample_tree["columns"]:
            self.sample_tree.heading(column, text=column)
        for row in rows:
            values = [row.get(column, "") for column in self.sample_tree["columns"]]
            self.sample_tree.insert("", "end", values=values)

    # ------------------------------------------------------------------
    def start_verification(self) -> None:
        if not self.loaded_rows:
            messagebox.showwarning("No data", "Please import a lead file before starting verification.")
            return

        mapping = self.mapping_frame.get_mapping()
        if "name" not in mapping or not mapping["name"]:
            messagebox.showwarning("Missing mapping", "Select at least the column that represents the full name.")
            return

        if self.current_task and not self.current_task.done():
            messagebox.showinfo("Job running", "A verification job is already in progress.")
            return

        self.result_rows.clear()
        self.refresh_result_table()
        self.progress_var.set(0.0)
        self.status_var.set("Starting verification...")
        self.source_styles.clear()
        self._color_cycle = itertools.cycle(self.SOURCE_COLORS)

        self._cancel_event = threading.Event()

        def progress_callback(current: int, total_rows: int) -> None:
            self.event_queue.put(("progress", current, total_rows))

        def result_callback(result: AggregatedLeadResult) -> None:
            self.event_queue.put(("result", result))

        def worker() -> List[AggregatedLeadResult]:
            try:
                results = run_verification_job(
                    self.orchestrator,
                    self.loaded_rows,
                    mapping,
                    cancel_event=self._cancel_event,
                    progress_callback=progress_callback,
                    result_callback=result_callback,
                )
            except Exception as exc:  # pragma: no cover - GUI surface
                self.event_queue.put(("error", exc))
                results = []
            cancelled = bool(self._cancel_event and self._cancel_event.is_set())
            self.event_queue.put(("done", results, cancelled))
            return results

        self.current_task = self._executor.submit(worker)
        self.status_var.set("Verification running...")

    # ------------------------------------------------------------------
    def cancel_verification(self) -> None:
        if self.current_task and not self.current_task.done():
            if self._cancel_event:
                self._cancel_event.set()
            self.status_var.set("Cancellation requested")

    # ------------------------------------------------------------------
    def clear_results(self) -> None:
        self.result_rows.clear()
        self.refresh_result_table()
        self.progress_var.set(0.0)
        self.status_var.set("Results cleared")
        self.source_styles.clear()
        self._color_cycle = itertools.cycle(self.SOURCE_COLORS)

    # ------------------------------------------------------------------
    def export_csv(self) -> None:
        if not self.result_rows:
            messagebox.showinfo("No results", "Run a verification job before exporting.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        try:
            write_results(path, self.result_rows)
        except Exception as exc:  # pragma: no cover - GUI surface
            messagebox.showerror("Export failed", str(exc))
            return
        messagebox.showinfo("Export complete", f"Results exported to {path}")

    # ------------------------------------------------------------------
    def export_excel(self) -> None:
        if not self.result_rows:
            messagebox.showinfo("No results", "Run a verification job before exporting.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Excel macro-enabled", "*.xlsm"),
                ("Excel template", "*.xltx *.xltm"),
            ],
        )
        if not path:
            return
        try:
            write_results(path, self.result_rows)
        except Exception as exc:  # pragma: no cover - GUI surface
            messagebox.showerror("Export failed", str(exc))
            return
        messagebox.showinfo("Export complete", f"Results exported to {path}")

    # ------------------------------------------------------------------
    def _poll_queue(self) -> None:
        try:
            while True:
                event = self.event_queue.get_nowait()
                self._handle_event(event)
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self._poll_queue)

    # ------------------------------------------------------------------
    def _handle_event(self, event: tuple) -> None:
        kind = event[0]
        if kind == "progress":
            _, current, total = event
            if total:
                self.progress_var.set(min(100.0, (current / total) * 100.0))
            else:
                self.progress_var.set(0.0)
            self.status_var.set(f"Processing lead {current} of {total}")
        elif kind == "result":
            _, result = event
            self.result_rows.append(result)
            self.refresh_result_table()
        elif kind == "error":
            _, exc = event
            messagebox.showerror("Verification failed", str(exc))
            self.status_var.set("Verification failed")
        elif kind == "done":
            _, results, cancelled = event
            if results:
                self.result_rows = results
            self.refresh_result_table()
            if cancelled:
                self.status_var.set("Verification cancelled")
            elif self.status_var.get() not in {"Verification failed", "Error starting verification"}:
                self.status_var.set("Verification finished")
            self.current_task = None
            self._cancel_event = None
            self.progress_var.set(100.0 if self.result_rows else 0.0)

    # ------------------------------------------------------------------
    def refresh_result_table(self) -> None:
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)

        term = self.filter_var.get().strip().lower()
        for result in self._filtered_results(term):
            sources = self._sources_for_result(result)
            tag = self._tag_for_sources(sources)
            self.results_tree.insert(
                "",
                "end",
                values=(
                    self._format_lead(result),
                    self._format_contacts(result),
                    ", ".join(sources) if sources else "—",
                    self._format_metadata(result),
                ),
                tags=(tag,),
            )

    # ------------------------------------------------------------------
    def _filtered_results(self, term: str) -> Iterable[AggregatedLeadResult]:
        if not term:
            return list(self.result_rows)
        filtered: List[AggregatedLeadResult] = []
        for result in self.result_rows:
            haystack = " ".join(
                filter(
                    None,
                    [
                        result.lead.display_name(),
                        result.lead.phone or "",
                        result.lead.email or "",
                        json.dumps(result.lead.metadata, ensure_ascii=False),
                        " ".join(contact.value for contact in result.contacts),
                        " ".join(
                            source
                            for contact in result.contacts
                            for source in contact.sources
                        ),
                    ],
                )
            ).lower()
            if term in haystack:
                filtered.append(result)
        return filtered

    # ------------------------------------------------------------------
    def _format_lead(self, result: AggregatedLeadResult) -> str:
        lead = result.lead
        parts = [lead.display_name()]
        if lead.phone:
            parts.append(lead.phone)
        if lead.email:
            parts.append(lead.email)
        location = ", ".join(filter(None, [lead.city, lead.state]))
        if location:
            parts.append(location)
        if lead.company:
            parts.append(lead.company)
        return " · ".join(filter(None, parts))

    # ------------------------------------------------------------------
    def _format_contacts(self, result: AggregatedLeadResult) -> str:
        if not result.contacts:
            return "No contacts found"
        return "; ".join(f"{contact.type}:{contact.value}" for contact in result.contacts)

    # ------------------------------------------------------------------
    def _format_metadata(self, result: AggregatedLeadResult) -> str:
        lead = result.lead
        metadata_parts = []
        if lead.company:
            metadata_parts.append(f"Company: {lead.company}")
        if lead.metadata:
            interesting = {k: v for k, v in lead.metadata.items() if k not in {"company", "city", "state", "name", "phone", "email"}}
            if interesting:
                metadata_parts.extend(f"{key}: {value}" for key, value in interesting.items())
        return "; ".join(metadata_parts) if metadata_parts else "—"

    # ------------------------------------------------------------------
    def _sources_for_result(self, result: AggregatedLeadResult) -> List[str]:
        sources = sorted({source for contact in result.contacts for source in contact.sources})
        if not sources and result.raw_results:
            sources = sorted({raw.source for raw in result.raw_results if getattr(raw, "source", "")})
        return sources

    # ------------------------------------------------------------------
    def _tag_for_sources(self, sources: Iterable[str]) -> str:
        key = ",".join(sources) if sources else "(unknown)"
        if key not in self.source_styles:
            style = next(self._color_cycle)
            self.source_styles[key] = style
            tag_name = f"source::{key}"
            self.results_tree.tag_configure(tag_name, background=style.background, foreground=style.foreground)
        else:
            style = self.source_styles[key]
            tag_name = f"source::{key}"
            self.results_tree.tag_configure(tag_name, background=style.background, foreground=style.foreground)
        return tag_name

    # ------------------------------------------------------------------
    def on_close(self) -> None:
        if self.current_task and not self.current_task.done():
            if not messagebox.askyesno("Quit", "A verification job is running. Quit anyway?"):
                return
            if self._cancel_event:
                self._cancel_event.set()
        self._executor.shutdown(wait=False, cancel_futures=True)
        for scraper in getattr(self.orchestrator, "scrapers", []):  # pragma: no cover - cleanup
            close = getattr(scraper, "close", None)
            if callable(close):
                try:
                    close()
                except Exception:  # pragma: no cover - defensive cleanup
                    LOGGER.exception("Failed to close scraper %s", scraper)
        self.root.destroy()

    # ------------------------------------------------------------------
    def _build_orchestrator(self) -> VerificationOrchestrator:
        config_path = os.environ.get("LEAD_VERIFIER_CONFIG")
        scrapers = []
        if config_path:
            try:
                config = load_configuration(config_path)
                scrapers = build_scrapers(config)
            except (ConfigurationError, FileNotFoundError, ValueError) as exc:
                messagebox.showwarning("Configuration error", f"Failed to load configuration: {exc}")
        if not scrapers:
            LOGGER.warning("No scrapers configured for UI - falling back to EchoScraper")
            scrapers = [EchoScraper()]
        return VerificationOrchestrator(scrapers)


def main() -> None:
    root = tk.Tk()
    app = LeadVerifierApp(root)
    root.mainloop()


if __name__ == "__main__":  # pragma: no cover - manual launch
    main()
