"""Input/output helpers for lead data and aggregated verification results."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable, List

from .models import AggregatedLeadResult, LeadInput


_CSV_SUFFIXES = {".csv"}
_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _normalise_key(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def load_leads(path: str | Path) -> List[LeadInput]:
    file_path = Path(path)
    if file_path.suffix.lower() in _CSV_SUFFIXES:
        return _load_leads_from_csv(file_path)
    if file_path.suffix.lower() in _EXCEL_SUFFIXES:
        return _load_leads_from_excel(file_path)
    raise ValueError(f"Unsupported input format '{file_path.suffix}'. Use CSV or Excel spreadsheet")


def _load_leads_from_csv(path: Path) -> List[LeadInput]:
    leads: List[LeadInput] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            leads.append(_row_to_lead(row))
    return leads


def _load_leads_from_excel(path: Path) -> List[LeadInput]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependency optional
        raise RuntimeError("Reading Excel files requires the 'openpyxl' package") from exc

    workbook = load_workbook(filename=path, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []
    header = [str(cell).strip() for cell in header_row if cell is not None]
    leads: List[LeadInput] = []
    for cells in rows:
        row = {header[idx]: value for idx, value in enumerate(cells) if idx < len(header)}
        leads.append(_row_to_lead(row))
    return leads


def _row_to_lead(row: dict) -> LeadInput:
    normalised = {_normalise_key(key): value for key, value in row.items() if key}

    first_name = normalised.get("first_name")
    last_name = normalised.get("last_name")
    name = normalised.get("name") or " ".join(filter(None, [first_name, last_name])) or None
    phone = normalised.get("phone")
    email = normalised.get("email")

    metadata = {key: value for key, value in row.items() if _normalise_key(key) not in {"first_name", "last_name", "name", "phone", "email"}}
    metadata.update(
        {
            key: value
            for key, value in {
                "city": normalised.get("city"),
                "state": normalised.get("state"),
                "address": normalised.get("address"),
            }.items()
            if value is not None and value != ""
        }
    )

    return LeadInput(
        name=name,
        phone=phone,
        email=email,
        first_name=first_name,
        last_name=last_name,
        metadata={k: str(v) if v is not None else "" for k, v in metadata.items()},
    )


def write_results(path: str | Path, results: Iterable[AggregatedLeadResult]) -> None:
    file_path = Path(path)
    if file_path.suffix.lower() in _CSV_SUFFIXES:
        _write_results_to_csv(file_path, results)
        return
    if file_path.suffix.lower() in _EXCEL_SUFFIXES:
        _write_results_to_excel(file_path, results)
        return
    raise ValueError(f"Unsupported output format '{file_path.suffix}'. Use CSV or Excel spreadsheet")


def _format_contacts(result: AggregatedLeadResult) -> str:
    formatted = []
    for contact in result.contacts:
        sources = ", ".join(contact.sources)
        formatted.append(f"{contact.type}:{contact.value} [{sources}]")
    return "; ".join(formatted)


def _write_results_to_csv(path: Path, results: Iterable[AggregatedLeadResult]) -> None:
    fieldnames = [
        "first_name",
        "last_name",
        "city",
        "state",
        "input_phone",
        "input_email",
        "contacts",
        "metadata",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow(
                {
                    "first_name": result.lead.first_name or "",
                    "last_name": result.lead.last_name or "",
                    "city": result.lead.city or "",
                    "state": result.lead.state or "",
                    "input_phone": result.lead.phone or "",
                    "input_email": result.lead.email or "",
                    "contacts": _format_contacts(result),
                    "metadata": json.dumps(result.lead.metadata, ensure_ascii=False),
                }
            )


def _write_results_to_excel(path: Path, results: Iterable[AggregatedLeadResult]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependency optional
        raise RuntimeError("Writing Excel files requires the 'openpyxl' package") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "first_name",
        "last_name",
        "city",
        "state",
        "input_phone",
        "input_email",
        "contacts",
        "metadata",
    ])
    for result in results:
        sheet.append(
            [
                result.lead.first_name or "",
                result.lead.last_name or "",
                result.lead.city or "",
                result.lead.state or "",
                result.lead.phone or "",
                result.lead.email or "",
                _format_contacts(result),
                json.dumps(result.lead.metadata, ensure_ascii=False),
            ]
        )
    workbook.save(path)
